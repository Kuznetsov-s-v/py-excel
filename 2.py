
from openpyxl.reader.excel import load_workbook
import openpyxl
book = load_workbook('res.xlsx')
# Определяем рабочий лист
ws = book.worksheets[0]
qtr = []
kol = []
str_x = []
kol_x = []
vr = []
delete_a = []
delete_b = []
# Определяем необходимую ячейку на листе
Cell_A = ws['A']
Cell_B = ws['B']
x = 0
for i in Cell_A:
    vr = [i.value]
    qtr.append(vr)
for j in Cell_B:
    vr = [j.value]
    kol.append(vr)

# список уникальных значенией "код дирекции"
j = 0
for i in qtr:
    if i not in str_x:
        str_x.append(qtr[j])
    j += 1

dct = {}
q = 0
for i in qtr:
    for j in str_x:
        if i == j:
            if i not in delete_a:
                delete_a.append(qtr[q])
                delete_b.append(kol[q])
            else:
                p = 0
                for u in delete_a:
                    if i == u:
                        delete_b[p].append(kol[q][0])
                    p += 1
    q += 1

for i in range(0,len(delete_a)):
    print(f'{delete_a[i]}    {delete_b[i]}')
def save():
    Poisk_w = openpyxl.Workbook()
    ws = Poisk_w.create_sheet('Итог')
    r = 1
    for statN in delete_a:
        ws.cell(row=r, column=1).value = statN[0]
        r += 1
    r = 1
    for statN in delete_b:
        ws.cell(row=r, column=2).value = " | ".join(statN)
        r += 1
    Poisk_w.save(filename='itog2.xlsx')
save()