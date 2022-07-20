from itertools import groupby

from openpyxl.reader.excel import load_workbook
import openpyxl
book = load_workbook('spr.xlsx')
# Определяем рабочий лист
ws = book.worksheets[0]
str = []
kol = []
str_x = []
kol_x = []
vr = []
Pustie_stroki = []
spisok_index_pust_str = []
# Определяем необходимую ячейку на листе
Cell_A = ws['A']
Cell_B = ws['B']
x = 0
for i in Cell_A:
    vr = [i.value]
    str.append(vr)
for j in Cell_B:
    vr = [j.value]
    kol.append(vr)
    # поиск пустых строк
    if vr[0] is None:
        Pustie_stroki.append(str[x])
    x += 1

# удаленеи совпадений для последующего поиска
Pustie_stroki = [el for el, _ in groupby(Pustie_stroki)]

Poisk_w = openpyxl.Workbook()
sh = Poisk_w['Sheet']
sh.title = 'Исключения'
# поиск повторов по пустым строкам
def find_N2():
    delete_a = []
    delete_b = []
    q = 0
    for i in str:
        for j in Pustie_stroki:
            if i == j:
                delete_a.append(str[q])
                delete_b.append(kol[q])
               # print(f'{delete_a[w]}    {delete_b[w]} \n')
                spisok_index_pust_str.append(q)
        q += 1
    sh.cell(row=1, column=1).value = 'Код дирекции (код по заявке)'
    sh.cell(row=1, column=2).value = 'колонка'
    r = 2
    for statN in delete_a:
        sh.cell(row=r, column=1).value = statN[0]
        r += 1
    r = 2
    for statN in delete_b:
        sh.cell(row=r, column=2).value = statN[0]
        r += 1
    #sh['A'] = delete_a
    #sh['B'] = delete_b

find_N2()

# удаление повторов c пробелом
def FindAndDelete2():
    spisok_index_pust_str.reverse()
    for i in spisok_index_pust_str:
        del str[i]
        del kol[i]
#удаление повторов с друг-другом
    j = 0
    for i in str:
        if i not in str_x:
            str_x.append(str[j])
            kol_x.append(kol[j])
        j += 1
FindAndDelete2()

def save():
    ws = Poisk_w.create_sheet('Итог')
    r = 1
    # если требуется удалить данные только совпадающие с пробелом, то in str: ,а если с пробелом и друг-другом, то str_x
    for statN in str_x:
        ws.cell(row=r, column=1).value = statN[0]
        r += 1
    r = 1
    for statN in kol_x:
        ws.cell(row=r, column=2).value = statN[0]
        r += 1
    # for z in range(0,len(str)):
    #    print(f'{str[z]}    {kol[z]} \n')

    wz = Poisk_w.create_sheet('Пустые колонки')
    wz.cell(row=1, column=1).value = 'Код дирекции (код по заявке)'
    wz.cell(row=1, column=2).value = '2 колонка'
    r = 2
    for statN in Pustie_stroki:
        wz.cell(row=r, column=1).value = statN[0]
        r += 1
    Poisk_w.save(filename='itog.xlsx')
save()